import openpyxl
import re
from openpyxl.utils import column_index_from_string
import numpy as np  # For NaN values
import pandas as pd


def extract_tables_with_column_names_and_dependencies(file_path):
    """
    Extracts tables with both computed values and metadata, including formulas, column names, and dependencies.
    """
    def extract_table_bounds(sheet):
        """Identify individual tables and combine those with the same column range."""
        table_ranges = []
        current_table_start = None
        non_empty_rows = []

        # Identify non-empty rows
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if any(cell is not None for cell in row):  # Check if row is not empty
                non_empty_rows.append(row_idx)
                if current_table_start is None:
                    current_table_start = row_idx
            elif current_table_start is not None:
                # End of a table
                table_ranges.append((current_table_start, row_idx - 1))
                current_table_start = None

        # Handle the last table if the sheet ends with data
        if current_table_start is not None:
            table_ranges.append((current_table_start, non_empty_rows[-1]))

        # Identify column ranges for each table and combine overlapping ones
        combined_tables = []
        for start_row, end_row in table_ranges:
            start_col = None
            end_col = None
            for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
                non_empty_cols = [idx for idx, cell in enumerate(row, start=1) if cell is not None]
                if non_empty_cols:
                    if start_col is None or min(non_empty_cols) < start_col:
                        start_col = min(non_empty_cols)
                    if end_col is None or max(non_empty_cols) > end_col:
                        end_col = max(non_empty_cols)

            # Combine tables with the same column range
            if combined_tables and combined_tables[-1][1] + 1 >= start_row and \
                    combined_tables[-1][2] == start_col and combined_tables[-1][3] == end_col:
                # Extend the row range of the previous table
                combined_tables[-1] = (combined_tables[-1][0], end_row, start_col, end_col)
            else:
                # Add a new table range
                combined_tables.append((start_row, end_row, start_col, end_col))

        for start_row, end_row, start_col, end_col in combined_tables:
            yield (start_row, start_col, end_row, end_col)

    def get_column_name(sheet, start_row, end_row, col_idx):
        """Determine the column name from the first non-None, non-formula cell in the column."""
        for row_idx in range(start_row, end_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.data_type != 'f' and cell.value is not None:
                return str(cell.value)
        return np.nan

    def extract_table_data(sheet, start_row, start_col, end_row, end_col):
        """Extract the actual table data."""
        table_data = []
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, 
                                   min_col=start_col, max_col=end_col, values_only=True):
            table_data.append(list(row))
        return table_data

    def extract_table_metadata(sheet, start_row, start_col, end_row, end_col, workbook):
        """Extract column metadata for the table."""
        column_metadata = []

        for col_idx in range(start_col, end_col + 1):
            column_formula = None
            dependency_sheets = set()

            for row_idx in range(start_row, end_row + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.data_type == 'f' and not column_formula:
                    column_formula = str(cell.value)
                    if "#REF!" in column_formula:
                        continue
                    for token in column_formula.split("!"):
                        for sheetname in workbook.sheetnames:
                            if sheetname in token.strip("="):
                                dependency_sheets.add(sheetname)

            column_name = get_column_name(sheet, start_row, end_row, col_idx)

            column_metadata.append({
                "ColumnName": column_name,
                "Formula": column_formula,
                "Dependencies": list(dependency_sheets),
            })

        return column_metadata

    workbook_values = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    workbook_formulas = openpyxl.load_workbook(file_path, data_only=False, read_only=False)

    visible_sheet_names = [sheet_name for sheet_name in workbook_values.sheetnames if not workbook_values[sheet_name].sheet_state == 'hidden']

    tables_by_sheet = {}
    for sheet_name in visible_sheet_names:
        sheet_values = workbook_values[sheet_name]
        sheet_formulas = workbook_formulas[sheet_name]
        tables = []

        for start_row, start_col, end_row, end_col in extract_table_bounds(sheet_values):
            # Extract table data
            table_data = extract_table_data(sheet_values, start_row, start_col, end_row, end_col)

            # Extract metadata
            column_metadata = extract_table_metadata(sheet_formulas, start_row, start_col, end_row, end_col, workbook_formulas)

            tables.append({
                "Coordinates": {"StartRow": start_row, "StartCol": start_col, "EndRow": end_row, "EndCol": end_col},
                "TableData": table_data,  # Store table data here
                "Metadata": column_metadata,
            })

        tables_by_sheet[sheet_name] = tables

    return tables_by_sheet

def create_cell_to_column_map(tables_with_metadata):
    """Create a global mapping of cells to column names."""
    cell_to_column_map = {}
    for sheet_name, tables in tables_with_metadata.items():
        for table in tables:
            coords = table["Coordinates"]
            metadata = table["Metadata"]
            for col_idx, meta in enumerate(metadata, start=coords["StartCol"]):
                for row_idx in range(coords["StartRow"], coords["EndRow"] + 1):
                    cell_to_column_map[(sheet_name, row_idx, col_idx)] = meta["ColumnName"]
    return cell_to_column_map

def enhance_formula_with_column_names(formula, cell_to_column_map, current_sheet_name):
    """Enhance formulas with column names, handling cross-sheet references."""
    cell_reference_pattern = re.compile(r"(?:(\w+)\!)?([A-Z]+)(\d+)")

    def replace_reference(match):
        sheet_name = match.group(1) or current_sheet_name
        col_letter = match.group(2)
        row_number = int(match.group(3))

        try:
            # Convert column letter to index
            col_index = column_index_from_string(col_letter)
        except ValueError:
            # If the column letter is invalid, leave it as-is
            return match.group(0)

        # Lookup in the precomputed map
        col_name = cell_to_column_map.get((sheet_name, row_number, col_index), "Unknown")

        return f"{match.group(0)}({col_name})"

    return cell_reference_pattern.sub(replace_reference, formula)

# Load and process the spreadsheet
file_path = '/Users/ajay/Documents/Atomic/inventory_analysis/Data/Company 1 - Inventory Planning.xlsx'
# file_path = '/Users/ajay/Documents/Atomic/inventory_analysis/Data/Company 2 - Supply Management.xlsx'
# file_path = '/Users/ajay/Documents/Atomic/inventory_analysis/Data/Company 3 - Inventory Dashboard _V2.xlsx'
tables_with_metadata = extract_tables_with_column_names_and_dependencies(file_path)
cell_to_column_map = create_cell_to_column_map(tables_with_metadata)

# Enhance formulas in all metadata
for sheet_name, tables in tables_with_metadata.items():
    for table in tables:
        for col_meta in table["Metadata"]:
            if col_meta["Formula"]:
                col_meta["EnhancedFormula"] = enhance_formula_with_column_names(
                    col_meta["Formula"], cell_to_column_map, sheet_name
                )


# Display the extracted tables and metadata
for sheet_name, tables in tables_with_metadata.items():
    print(f"Sheet: {sheet_name}\n{'-' * 40}")
    
    for i, table_info in enumerate(tables):
        print(f"Sheet: {sheet_name}\n{'-' * 40}")
        print(f"  Table {i + 1}:\n{'-' * 20}")
        
        # Safely extract and display table data if it exists
        table_data = table_info.get("TableData", None)
        if table_data:
            print("    Table Data (First 6 Rows):")
            df = pd.DataFrame(table_data)
            print(df.head(6))  # Show the first 6 rows as a DataFrame
        else:
            print("    Table Data: Not available")
        
        # Safely extract and display metadata line by line if it exists
        metadata = table_info.get("Metadata", None)
        if metadata:
            print("    Metadata:")
            for col_meta in metadata:
                print(f"      {col_meta}")
        else:
            print("    Metadata: Not available")
        
        print("\n")
