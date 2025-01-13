import os
import traceback
from langchain.chat_models import ChatOpenAI
from langchain_experimental.tools import PythonREPLTool
from openpyxl import load_workbook
import re



# Function to extract a 50xN chunk from a sheet
def extract_sheet_chunk(sheet_name):
    sheet_name = sheet_name.strip().replace('"', '').replace("'", "")
    if sheet_name not in sheet_names:
        log_message = f"Sheet '{sheet_name}' doesn't exist. Available sheets: {sheet_names}"
        print(log_message)
        return log_message
    print(f"Fetching 50xN chunk from sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    max_cols = sheet.max_column
    data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=1, max_row=50, max_col=max_cols)]
    formatted_chunk = "\n".join([", ".join(map(str, row)) for row in data])
    print(f"Fetched chunk from sheet '{sheet_name}':\n{formatted_chunk[:500]}...")  # Show first 500 chars for brevity
    return formatted_chunk

# Function to analyze a chunk and generate code


def analyze_and_generate_code(sheet_name: str, chunk: str, sheet_folder: str) -> str:
    llm_input = f"""
    You are a Python code generator for table extraction.
    Below is a 50xN chunk of data from the sheet '{sheet_name}' in an Excel file.

    {chunk}

    Tasks:
    1. Analyze the data to identify all tables present in the chunk. Tables are defined as contiguous blocks of data separated by blank rows or rows with no data.
    2. Generate Python code that:
       - Reads data from the provided Excel file ('{file_path}') and specified sheet ('{sheet_name}').
       - Saves all extracted tables as separate CSV files in the folder '{sheet_folder}'.
       - Ensures the generated code creates the CSV files in the exact folder specified and avoids creating additional folders elsewhere.
    3. Do not make any assumptions about the meaning or functionality of specific columns unless explicitly mentioned.
    """
    print(f"Sending data to LLM for analysis and code generation...")
    llm = ChatOpenAI(model="gpt-4o-mini", temperature=0)
    try:
        response = llm.invoke(llm_input)
        generated_response = response.content  # Extract string content
        print(f"Full LLM response:\n{generated_response[:500]}...")  # Log full response

        # Extract Python code block using regex
        match = re.search(r"```python(.*?)```", generated_response, re.DOTALL)
        if match:
            generated_code = match.group(1).strip()  # Extract the Python code
            print(f"Extracted code:\n{generated_code[:500]}...")  # Log extracted code
            return generated_code
        else:
            print("No valid Python code block found in the LLM response.")
            return "Error: No valid Python code block found."
    except Exception as e:
        print(f"Error during code generation: {e}")
        return f"Error during code generation: {e}"

# Function to execute generated Python code
def execute_code(code: str) -> str:
    try:
        print(f"Executing the following code:\n{code}...")  # Log the extracted code
        repl_tool = PythonREPLTool()
        result = repl_tool.run(code)
        print("Execution result:")
        print(result)
        return result
    except Exception as e:
        error_message = f"Execution Error: {str(e)}"
        print(error_message)
        print(traceback.format_exc())
        return error_message

# Main execution process
def run_analysis(sheet_name):
    print(f"Starting analysis for sheet: {sheet_name}")
    
    # Step 1: Fetch sheet chunk
    chunk = extract_sheet_chunk(sheet_name)
    if "doesn't exist" in chunk:
        print(f"Error: {chunk}")
        return chunk

    # Step 2: Create the sheet folder
    sheet_folder = os.path.join(base_dir, sheet_name)
    os.makedirs(sheet_folder, exist_ok=True)

    # Step 3: Analyze and generate code
    print("Analyzing the data and generating Python code...")
    code = analyze_and_generate_code(sheet_name, chunk, sheet_folder)
    if "Error during code generation" in code:
        return code

    # Step 4: Attempt to execute code with retries
    retry = 10
    while retry > 0:
        result = execute_code(code)
        if "Execution Error" not in result:
            print("Code executed successfully!")
            return result  # Successful execution

        # Provide feedback and request fixes
        error_message = f"\n\nError Message: {result}"
        print(f"Retrying... ({10 - retry + 1}/10)")
        print(f"Requesting LLM to fix the code. Error: {error_message}")
        code = analyze_and_generate_code(sheet_name, chunk, sheet_folder)
        retry -= 1

    print("Exceeded maximum retries. Could not execute code successfully.")
    return "Failed after multiple retries."

for file_path in ['/Users/ajay/Documents/Atomic/inventory_analysis/Data/Company 1 - Inventory Planning.xlsx', '/Users/ajay/Documents/Atomic/inventory_analysis/Data/Company 2 - Supply Management.xlsx', '/Users/ajay/Documents/Atomic/inventory_analysis/Data/Company 3 - Inventory Dashboard _V2.xlsx']:
    print("Processing File",file_path)
    # Load Excel Workbook
    # file_path = '/Users/ajay/Documents/Atomic/inventory_analysis/Data/Company 3 - Inventory Dashboard _V2.xlsx'
    workbook_name = os.path.splitext(os.path.basename(file_path))[0]  # Get workbook name without extension
    workbook = load_workbook(file_path, data_only=False, read_only=True)
    sheet_names = [sheet_name for sheet_name in workbook.sheetnames if not workbook[sheet_name].sheet_state == 'hidden']

    # Create a base directory with the workbook name
    base_dir = os.path.join(os.getcwd(), workbook_name)
    os.makedirs(base_dir, exist_ok=True)

    # Run the workflow for each sheet
    for sheet_to_analyze in sheet_names:
        print('-'*40, f'\nAnalyzing Sheet: {sheet_to_analyze}')
        print(run_analysis(sheet_to_analyze))
